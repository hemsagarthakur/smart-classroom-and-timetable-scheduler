import os
from collections import Counter
from datetime import datetime

from bson import ObjectId
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from database.db import mongo


DEPT_COLORS = {
    "CSE": "D8F3DC",
    "ECE": "E3F2FD",
    "MECH": "FFF3CD",
    "MBA": "FCE4EC",
}


def _to_object_id(value):
    return value if isinstance(value, ObjectId) else ObjectId(value)


def export_to_excel(variant_doc):
    workbook = Workbook()
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    entries = variant_doc.get("entries", [])
    periods = sorted({entry["period_number"] for entry in entries})
    faculty_stats = {}
    room_usage = Counter(entry["room_code"] for entry in entries)
    faculty_docs = {
        str(item["_id"]): item
        for item in mongo.db.faculty.find({"_id": {"$in": [_to_object_id(entry["faculty_id"]) for entry in entries]}})
    }
    room_docs = {
        str(item["_id"]): item
        for item in mongo.db.rooms.find({"_id": {"$in": [_to_object_id(entry["room_id"]) for entry in entries]}})
    }

    ws = workbook.active
    ws.title = "Timetable"
    ws.cell(row=1, column=1, value="Period / Time").font = Font(bold=True)
    for idx, day in enumerate(headers, start=2):
        cell = ws.cell(row=1, column=idx, value=day)
        cell.font = Font(bold=True)
        cell.border = border

    for row_idx, period in enumerate(periods, start=2):
        period_entries = [entry for entry in entries if entry["period_number"] == period]
        label = f"Period {period}"
        if period_entries:
            label = f"Period {period} ({period_entries[0]['start_time']}-{period_entries[0]['end_time']})"
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=1).border = border

        for col_idx, day in enumerate(headers, start=2):
            entry = next((item for item in entries if item["period_number"] == period and item["day"] == day), None)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if entry:
                cell.value = f"{entry['subject_name']}\n{entry['faculty_name']}\n{entry['room_code']}"
                color = DEPT_COLORS.get(entry.get("department_code"), "F8F9FA")
                cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
                if entry["faculty_id"] not in faculty_stats:
                    faculty_doc = faculty_docs.get(str(entry["faculty_id"]), {})
                    faculty_stats[entry["faculty_id"]] = {
                        "name": entry["faculty_name"],
                        "email": faculty_doc.get("email", ""),
                        "dept": entry.get("department_code", ""),
                        "hours": 0,
                        "max_hours": faculty_doc.get("max_hours_per_week", 20),
                    }
                faculty_stats[entry["faculty_id"]]["hours"] += 1
            else:
                cell.value = "-"
                cell.fill = PatternFill(fill_type="solid", start_color="EEEEEE", end_color="EEEEEE")

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 28)

    workload_ws = workbook.create_sheet("Faculty Workload")
    workload_ws.append(["Faculty Name", "Email", "Dept", "Total Hours", "Max Hours", "Status"])
    for cell in workload_ws[1]:
        cell.font = Font(bold=True)
        cell.border = border
    for stat in faculty_stats.values():
        status = "Over Limit" if stat["hours"] > stat["max_hours"] else "Within Limit"
        workload_ws.append([stat["name"], stat["email"], stat["dept"], stat["hours"], stat["max_hours"], status])
    for row in workload_ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
        if row[5].value == "Over Limit":
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", start_color="F8D7DA", end_color="F8D7DA")

    room_ws = workbook.create_sheet("Room Utilization")
    room_ws.append(["Room Code", "Type", "Capacity", "Total Slots Used", "Utilization %"])
    for cell in room_ws[1]:
        cell.font = Font(bold=True)
        cell.border = border
    total_periods = max(len(periods) * len(headers), 1)
    room_meta = {}
    for entry in entries:
        room_doc = room_docs.get(str(entry["room_id"]), {})
        room_meta.setdefault(
            entry["room_code"],
            {"type": room_doc.get("room_type", "classroom"), "capacity": room_doc.get("capacity", "-"), "used": 0},
        )
        room_meta[entry["room_code"]]["used"] += 1
    for room_code, meta in room_meta.items():
        utilization = round((meta["used"] / total_periods) * 100, 2)
        room_ws.append([room_code, meta["type"], meta["capacity"], meta["used"], utilization])
    for row in room_ws.iter_rows():
        for cell in row:
            cell.border = border

    export_dir = os.path.join(current_app.root_path, current_app.config["EXPORT_FOLDER"])
    os.makedirs(export_dir, exist_ok=True)
    filename = f"timetable_{variant_doc['_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(export_dir, filename)
    workbook.save(file_path)
    return file_path
